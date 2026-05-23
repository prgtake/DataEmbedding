# =====================================================
#  Test Data Generator for Data Embedding Macro
#  Copyright (c) 2026 Datan (データン)
#  Licensed under the MIT License.
# =====================================================
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Color
from openpyxl.worksheet.datavalidation import DataValidation

# --- 1. マスターデータの作成 ---
data = {
    '社員番号': [f'EMP{i:04d}' for i in range(1, 21)],
    '氏名': [f'佐藤 健太{i}' if i%2==0 else f'高橋 明美{i}' for i in range(1, 21)],
    '所属部署': ['デジタル戦略部', '経営企画室', '人事総務グループ', 'グローバル営業部', '製品開発チーム'] * 4,
    '役職・等級': ['M2 (マネージャー)', 'S3 (シニア)', 'S1 (ジュニア)', 'M1 (リーダー)', 'S2 (プロ)'] * 4,
    '直近評価': ['A', 'A', 'B', 'S', 'B'] * 4,
    '期待される役割': [
        '次世代リーダーとしてチームを牽引',
        '専門性を活かしたプロジェクト推進',
        '基礎スキルの習得と業務完遂',
        '新規市場の開拓と海外拠点連携',
        '品質管理プロセスの標準化'
    ] * 4,
    '返信期限': ['2026/06/15'] * 20
}
df = pd.DataFrame(data)

# --- 2. Excelファイルの作成 ---
file_name = 'survey_data.xlsx'
writer = pd.ExcelWriter(file_name, engine='openpyxl')
df.to_excel(writer, sheet_name='データ', index=False)
workbook = writer.book

# --- 3. 雛形（入力シート）の作成 ---
ws = workbook.create_sheet('雛形')
ws.sheet_view.showGridLines = False

# スタイルの定義
accent_color = '4472C4' # ロイヤルブルー
light_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
header_fill = PatternFill(start_color=accent_color, end_color=accent_color, fill_type='solid')
input_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid') # 入力エリアは白

border_full = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
border_header = Border(bottom=Side(style='medium', color=accent_color))

# 列幅
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 35

# --- タイトルエリア ---
ws.merge_cells('B2:E3')
ws['B2'] = '2026年度 キャリア開発・自己申告シート'
ws['B2'].font = Font(name='Meiryo UI', size=20, bold=True, color='FFFFFF')
ws['B2'].fill = header_fill
ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

# --- 基本情報（マクロで自動埋め込み） ---
# 背景をグレーにして「ここは修正不要」感を出す
for r in range(5, 9):
    for c in ['B', 'C', 'D', 'E']:
        ws[f'{c}{r}'].fill = light_fill
        ws[f'{c}{r}'].font = Font(name='Meiryo UI', size=10)

ws['B5'] = '社員番号'
ws['C5'] = "='データ'!A2"
ws['D5'] = '氏名'
ws['E5'] = "='データ'!B2"

ws['B6'] = '所属部署'
ws['C6'] = "='データ'!C2"
ws['D6'] = '役職・等級'
ws['E6'] = "='データ'!D2"

ws['B7'] = '直近の評価'
ws['C7'] = "='データ'!E2"
ws['D7'] = '返信期限'
ws['E7'] = "='データ'!G2"
ws['E7'].font = Font(name='Meiryo UI', size=10, color='FF0000', bold=True)

ws['B8'] = '期待される役割'
ws.merge_cells('C8:E8')
ws['C8'] = "='データ'!F2"
ws['C8'].alignment = Alignment(wrap_text=True)

# 罫線引き
for r in range(5, 9):
    for c in ['B', 'C', 'D', 'E']:
        ws[f'{c}{r}'].border = border_full

# --- 入力エリアの案内 ---
ws.merge_cells('B10:E10')
ws['B10'] = '▼ 以下、ご自身の考えを記入してください'
ws['B10'].font = Font(name='Meiryo UI', size=11, bold=True, color=accent_color)
ws['B10'].border = border_header

# --- セクション1: 本年度の振り返り ---
ws.merge_cells('B12:E12')
ws['B12'] = ' 1. 本年度の成果と課題（自己リフレクション）'
ws['B12'].font = Font(name='Meiryo UI', size=11, bold=True, color='FFFFFF')
ws['B12'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid') # グリーン

ws.merge_cells('B13:E16')
ws['B13'] = 'ここに具体的な成果と、もっと伸ばしたい点を記入してください...'
ws['B13'].font = Font(name='Meiryo UI', size=10, color='A6A6A6')
ws['B13'].alignment = Alignment(vertical='top', wrap_text=True)
for r in range(13, 17):
    for c in ['B', 'C', 'D', 'E']:
        ws[f'{c}{r}'].border = border_full

# --- セクション2: 次年度の目標 ---
ws.merge_cells('B18:E18')
ws['B18'] = ' 2. 次年度に挑戦したいこと・習得したいスキル'
ws['B18'].font = Font(name='Meiryo UI', size=11, bold=True, color='FFFFFF')
ws['B18'].fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid') # オレンジ

ws['B19'] = '優先度'
ws['C19'] = '具体的な目標'
ws['D19'] = '必要なリソース(研修等)'
ws['E19'] = '期限'
for c in ['B', 'C', 'D', 'E']:
    ws[f'{c}19'].fill = light_fill
    ws[f'{c}19'].font = Font(name='Meiryo UI', size=10, bold=True)
    ws[f'{c}19'].alignment = Alignment(horizontal='center')
    ws[f'{c}19'].border = border_full

for r in range(20, 23):
    for c in ['B', 'C', 'D', 'E']:
        ws[f'{c}{r}'].border = border_full

# プルダウンの追加（優先度）
dv = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
ws.add_data_validation(dv)
dv.add('B20:B22')

# --- セクション3: 会社への要望 ---
ws.merge_cells('B24:E24')
ws['B24'] = ' 3. 会社・上司への要望・相談事項'
ws['B24'].font = Font(name='Meiryo UI', size=11, bold=True, color='FFFFFF')
ws['B24'].fill = header_fill

ws.merge_cells('B25:E27')
ws['B25'].border = border_full
for r in range(25, 28):
    for c in ['B', 'C', 'D', 'E']:
        ws[f'{c}{r}'].border = border_full

# 最後にフッター
ws['E29'] = '以上'
ws['E29'].alignment = Alignment(horizontal='right')

writer.close()
print("survey_data.xlsx has been created successfully.")
