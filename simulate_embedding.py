import openpyxl
import shutil
import os

# --- Settings ---
source_file = 'TESTデータ/survey_data.xlsx'
output_dir = 'output_results'
data_sheet_name = 'データ'
template_sheet_name = '雛形'

if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Load the source to get data
wb_source = openpyxl.load_workbook(source_file, data_only=False)
ws_data = wb_source[data_sheet_name]

# Get headers
headers = [cell.value for cell in ws_data[1]]
# Get data records (limit to 3 for demo)
records = []
for row in ws_data.iter_rows(min_row=2, max_row=4, values_only=True):
    if any(row):
        records.append(row)

wb_source.close()

print(f"Starting simulation for {len(records)} records...")

for record in records:
    emp_id = record[0]
    output_file = os.path.join(output_dir, f"{emp_id}.xlsx")
    
    # 1. Subtraction Method: Copy the whole file to preserve properties/styles
    shutil.copyfile(source_file, output_file)
    
    # 2. Open the copy
    wb = openpyxl.load_workbook(output_file)
    ws_data_out = wb[data_sheet_name]
    ws_temp_out = wb[template_sheet_name]
    
    # 3. Injection: Write the record data to the second row of the data sheet
    # (Assuming the template formulas refer to row 2)
    for col_idx, value in enumerate(record, 1):
        ws_data_out.cell(row=2, column=col_idx, value=value)
    
    # 4. Freeze Values: In a real Excel environment, formulas would update.
    # Here, we manually resolve the simple formulas used in the generator.
    # Generator uses: ='データ'!A2, ='データ'!B2, etc.
    # We'll just replace them with the values from the record.
    mapping = {
        "='データ'!A2": record[0],
        "='データ'!B2": record[1],
        "='データ'!C2": record[2],
        "='データ'!D2": record[3],
        "='データ'!E2": record[4],
        "='データ'!G2": record[6],
        "='データ'!F2": record[5]
    }
    
    for row in ws_temp_out.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in mapping:
                cell.value = mapping[cell.value]
    
    # 5. Subtract: Remove the 'Data' sheet
    del wb[data_sheet_name]
    
    # Rename template to EMP ID
    ws_temp_out.title = emp_id
    
    # 6. Save
    wb.save(output_file)
    print(f"Generated: {output_file}")

print("Simulation complete. Please check the 'output_results' folder.")
